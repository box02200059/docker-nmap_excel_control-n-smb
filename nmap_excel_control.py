#!/usr/bin/env python
# -*- coding: utf-8 -*-
import nmap
import time
from openpyxl import Workbook
f = open('/shared/target.txt', 'r')
target_list = []
lines = f.readlines()
for line in lines:
    if '/24' in line:
        for a in range(1, 255):
            target_list.append(line.rstrip('\r\n')[0:-4]+str(a))
    else:
        try:
            target_list.append(line.rstrip('\r\n'))
        except:
            target_list.append(line)
print(target_list)
#target_list = ['127.0.0.1']
f.close()

wb = Workbook()
ws = wb.active
ws['A1'] = 'ip_address'
ws['B1'] = 'port'
ws['C1'] = 'product'
ws['D1'] = 'version'
ws['E1'] = 'name'
ws['F1'] = 'cpe'
ws['G1'] = 'extrainfo'

num = 0
for target in target_list:
    num += 1
    num2 = 0
    print 'scanning %s' % target
    try:
        nm = nmap.PortScanner()
        nm.scan(target, '1-1000,1433,3306,3389,8080-8090','-sS -Av')
        print nm.command_line()
        for port in range(1, 10000):
            try:
                s1 = nm[target]['tcp'][port]['product']
                s2 = nm[target]['tcp'][port]['version']
                s3 = nm[target]['tcp'][port]['name']
                s4 = nm[target]['tcp'][port]['cpe']
                s5 = nm[target]['tcp'][port]['extrainfo']
                print target, port, s1, s2, s3, s4, s5
                ws.append([target, port, s1, s2, s3, s4, s5])
                num2 += 1
            except:
                pass
        if num2 == 0:
            num -= 1
    except:
        num -= 1
        print 'scanning %s error!' % target
    if num == 50:
        num = 0
        wb.save("/shared/%s.xlsx" % time.asctime(time.localtime(time.time())).replace(' ', '_').replace(':', '.')
                )
        print 'Save file.
if num > 0:
    wb.save("/shared/%s.xlsx" % time.asctime(time.localtime(time.time())).replace(' ', '_').replace(':', '.')
            )
    print 'Save last file.
